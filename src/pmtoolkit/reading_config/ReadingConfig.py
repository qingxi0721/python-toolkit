import configparser
import csv
import json
import os
from typing import Any, Literal

import openpyxl
import xlrd
import xmltodict
import yaml


class ReadingConfig:
    def get(self, input_file: str) -> Any:
        """获取config文件内容

        Args:
            input_file (str): 需要读取的文件路径

        Raises:
            ValueError: 当文件不是.ini/.json/.xml/.yaml的一种时抛出异常

        Returns:
            Any: 依据配置文件类型不同返回字典/字典数组形式
        """
        try:
            os.path.exists(input_file)
        except FileNotFoundError:
            return None
        else:
            file_type = input_file.split('.')[-1]
            # 读取ini
            if file_type == 'ini':
                return self._get_ini(input_file)

            # 读取json
            elif file_type == 'json':
                return self._get_json(input_file)

            # 读取yaml
            elif file_type == 'yaml':
                return self._get_yaml(input_file)

            # 读取xml
            elif file_type == 'xml':
                return self._get_xml(input_file)

            # 读取xlsx
            elif file_type == 'xlsx':
                return self._get_xlsx_to_dict(input_file)

            else:
                raise ValueError(f"Unsupported file type: {file_type}")

    def get_xlsx(self, input_file: str, return_type: Literal['dict', 'array'], header=None) -> Any:
        if return_type == 'dict':
            return self._get_xlsx_to_dict(input_file, header)
        elif return_type == 'array':
            return self._get_xlsx_to_array(input_file, header)
        else:
            raise ValueError(f"Unsupported return type: {return_type}")

    def get_csv(self, input_file: str, return_type: Literal['dict', 'array'], header=None) -> Any:
        if return_type == 'dict':
            return self._get_csv_to_dict(input_file)
        elif return_type == 'array':
            return self._get_csv_to_array(input_file, header)
        else:
            raise ValueError(f"Unsupported return type: {return_type}")

    @staticmethod
    def _get_ini(input_file: str) -> Any:
        """读取ini配置

        Args:
            input_file (str): 需读取文件

        Returns:
           Any: 返回字典类型
        """
        # 创建配置解析器 区分大小写
        config = configparser.RawConfigParser()
        config.optionxform = lambda option: option
        config.read(input_file)

        # 递归返回字典类型
        return {section: dict(config.items(section)) for section in config.sections()}

    @staticmethod
    def _get_json(input_file: str) -> Any:
        """读取json配置

        Args:
            input_file (str): 同_get_ini_

        Returns:
            Any: 返回字典/字典数组类型
        """
        with open(input_file, 'r') as f:
            return json.load(f)

    @staticmethod
    def _get_yaml(input_file: str) -> Any:
        """读取yaml配置

        Args:
            input_file (str): 同_get_ini_

        Returns:
            Any: 返回字典类型
        """
        with open(input_file, 'r') as f:
            return yaml.load(f, Loader=yaml.FullLoader)

    @staticmethod
    def _get_xml(input_file: str) -> Any:
        """读取xml配置

        Args:
            input_file (str): 同_get_ini_

        Returns:
            Any: 返回字典类型
        """
        with open(input_file) as f:
            return xmltodict.parse(f.read())

    @staticmethod
    def _get_xlsx_to_array(input_file: str, header=1) -> Any:
        def convert_to_array(excel_data: Any) -> Any:
            result = []
            for row_index, row in enumerate(excel_data.iter_rows(min_row=header + 1)):
                row_data = [cell.value for cell in row]
                result.append(row_data)
            return result

        file_type = input_file.split('.')[-1]
        if file_type == 'xlsx':
            workbook = openpyxl.load_workbook(input_file)
            sheet = workbook.active
            return convert_to_array(sheet)

        elif file_type == 'xls':
            workbook = xlrd.open_workbook(input_file)
            sheet = workbook.sheet_by_index(0)
            return convert_to_array(sheet)
        else:
            raise ValueError(f"Unsupported file type: {file_type}")

    @staticmethod
    def _get_xlsx_to_dict(input_file: str, header=1) -> Any:
        def convert_to_dict(excel_data: Any) -> list:
            result = []
            headers = [cell.value for cell in excel_data[header]]
            for row in excel_data.iter_rows(min_row=header + 1):
                row_data = [cell.value for cell in row]
                result.append(dict(zip(headers, row_data)))
            return result

        file_type = input_file.split('.')[-1]
        if file_type == 'xlsx':
            workbook = openpyxl.load_workbook(input_file)
            sheet = workbook.active
            return convert_to_dict(sheet)

        elif file_type == 'xls':
            workbook = xlrd.open_workbook(input_file)
            sheet = workbook.sheet_by_index(0)
            return convert_to_dict(sheet)
        else:
            raise ValueError(f"Unsupported file type: {file_type}")

    @staticmethod
    def _get_csv_to_dict(input_file: str) -> Any:
        with open(input_file, mode='r', newline='', encoding='utf-8') as file:
            reader = csv.DictReader(file)
            data = [row for row in reader]
        return data

    @staticmethod
    def _get_csv_to_array(input_file: str, header=1) -> Any:
        with open(input_file, mode='r', newline='', encoding='utf-8') as file:
            reader = csv.reader(file)
            result = []
            for row_index, row in enumerate(reader):
                if row_index + 1 == header:
                    continue
                result.append(row)
        return result
